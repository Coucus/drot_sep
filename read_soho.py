'''import pandas as pd
import matplotlib.pyplot as plt
with open('/Users/duan/Desktop/eph_2010_225_00_2010_228_00_30_1725532643.dat', 'r') as file:
    lines = file.readlines()
    first_data_line = 0
    for i, line in enumerate(lines):
        if line.strip() != 'julian date' in line.lower():  # 假设数据开始前有一个空行或特定字符串
            first_data_line = i + 1
            break
df = pd.read_csv('/Users/duan/Desktop/eph_2010_225_00_2010_228_00_30_1725532643.dat', sep='\s+', skiprows=first_data_line-1,header=1)
fig=plt.figure()
print(df)
plt.plot(df.iloc[:,2],df.iloc[:,8])
plt.show()'''
import pyspedas
import pytplot
from pytplot import tplot
from openpyxl import load_workbook
# 获取数据

erne_vars = pyspedas.soho.erne(trange=['2012-01-23/00:00:00', '2012-01-26/00:00:00'],datatype='hed_l2-1min', time_clip=True)

# 假设你已经有了一个名为 'my_excel_file.xlsx' 的 Excel 文件
file_name = '/Users/duan/Desktop/sep/class_m.xlsx'
sheet_name = 'Sheet1'  # 你想要写入的工作表名称

# 加载现有的 Excel 文件
wb = load_workbook(filename=file_name)
ws = wb[sheet_name]

# 获取 pytplot.data_quants['PH'].values 数据
# 并且你有10个通道的数据
for i in range(10):
    max_value = pytplot.data_quants['PH'].values[:, i].max()
    print(max_value)
    # 将数据写入第二行的第 i+1 列，因为列索引从 1 开始
    ws[f'{chr(68+i)}24'] = max_value

# 保存 Excel 文件
print('sucessful load')
wb.save(file_name)
Mev14_5=pytplot.data_quants['PH'].values[:,0]
Mev18=pytplot.data_quants['PH'].values[:,1]
Mev22_5=pytplot.data_quants['PH'].values[:,2]
Mev28_5=pytplot.data_quants['PH'].values[:,3]
Mev36=pytplot.data_quants['PH'].values[:,4]
Mev45=pytplot.data_quants['PH'].values[:,5]
Mev57=pytplot.data_quants['PH'].values[:,6]
Mev72=pytplot.data_quants['PH'].values[:,7]
Mev90=pytplot.data_quants['PH'].values[:,8]
Mev115=pytplot.data_quants['PH'].values[:,9]
time=pytplot.data_quants['PH'].coords['time'].values
time=pytplot.data_quants['PH'].coords['time'].values
pytplot.store_data("Mev14_5", data={'x':time, 'y':Mev14_5})
pytplot.store_data("Mev18", data={'x':time, 'y':Mev18})
pytplot.store_data("Mev22_5", data={'x':time, 'y':Mev22_5})
pytplot.store_data("Mev28_5", data={'x':time, 'y':Mev28_5})
pytplot.store_data("Mev36", data={'x':time, 'y':Mev36})
pytplot.store_data("Mev45", data={'x':time, 'y':Mev45})
pytplot.store_data("Mev57", data={'x':time, 'y':Mev57})
pytplot.store_data("Mev72", data={'x':time, 'y':Mev72})
pytplot.store_data("Mev90", data={'x':time, 'y':Mev90})
pytplot.store_data("Mev115", data={'x':time, 'y':Mev115})

tplot(['Mev14_5','Mev18','Mev22_5','Mev28_5','Mev36','Mev45','Mev57','Mev72','Mev90','Mev115'])
tplot(['PH'])
